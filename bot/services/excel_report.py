from __future__ import annotations

import io
from datetime import datetime, timedelta, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

_DISPLAY_TZ = timezone(timedelta(hours=3))

_HEADER_FONT = Font(bold=True)
_TITLE_FONT = Font(bold=True, size=14)
_CENTER = Alignment(horizontal="center")


def build_official_report_xlsx(report: dict) -> bytes:
    """Renders a plain official team report as an Excel workbook (one sheet per section)."""
    team_name = report.get("teamName") or "Команда"
    totals = report.get("totals") or {}
    members = report.get("members") or []
    weekday = report.get("weekdayProductivity") or []
    best_performer = report.get("bestPerformer")
    excuse_stats = report.get("excuseStats") or []

    wb = Workbook()

    summary = wb.active
    summary.title = "Общая статистика"
    summary["A1"] = f"Отчёт по команде «{team_name}»"
    summary["A1"].font = _TITLE_FONT
    summary["A2"] = f"Сформирован: {_format_generated_at(report.get('generatedAt'))}"
    summary.append([])
    _write_table(summary, ["Показатель", "Значение"], [
        ["Всего задач", totals.get("totalTasks", 0)],
        ["Закрыто", totals.get("completedTasks", 0)],
        ["В работе", totals.get("activeTasks", 0)],
        ["Просрочено", totals.get("overdueTasks", 0)],
    ])
    summary.append([])
    summary.append(["Лучший работник"])
    summary.cell(row=summary.max_row, column=1).font = _HEADER_FONT
    if best_performer:
        summary.append([best_performer.get("username", "—"), f"{best_performer.get('completedTasks', 0)} задач закрыто"])
    else:
        summary.append(["Нет данных"])
    _autosize(summary)

    members_sheet = wb.create_sheet("По участникам")
    member_rows = [
        [
            m.get("username", "—"),
            m.get("totalTasks", 0),
            m.get("completedTasks", 0),
            m.get("activeTasks", 0),
            m.get("overdueTasks", 0),
        ]
        for m in members
    ]
    _write_table(members_sheet, ["Участник", "Всего", "Закрыто", "В работе", "Просрочено"], member_rows)
    _autosize(members_sheet)

    weekday_sheet = wb.create_sheet("Дни недели")
    _write_table(weekday_sheet, ["День недели", "Закрыто задач"], [[d.get("day", ""), d.get("count", 0)] for d in weekday])
    _autosize(weekday_sheet)

    excuse_sheet = wb.create_sheet("Больничные за месяц")
    _write_table(excuse_sheet, ["Участник", "Больничные/отгулы"], [[s.get("username", "—"), s.get("count", 0)] for s in excuse_stats])
    _autosize(excuse_sheet)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_table(sheet, headers: list[str], rows: list[list]) -> None:
    start_row = sheet.max_row + 1 if sheet.max_row > 1 or sheet["A1"].value else 1
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=start_row, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
    if not rows:
        sheet.cell(row=start_row + 1, column=1, value="Нет данных")
        return
    for row in rows:
        sheet.append(row)


def _autosize(sheet) -> None:
    for column_cells in sheet.columns:
        length = max((len(str(cell.value)) for cell in column_cells if cell.value is not None), default=0)
        sheet.column_dimensions[column_cells[0].column_letter].width = max(12, length + 2)


def _format_generated_at(value: str | None) -> str:
    if not value:
        return "—"
    try:
        dt = datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError:
        return value
    return dt.astimezone(_DISPLAY_TZ).strftime("%d.%m.%Y %H:%M")
